# -*- coding: utf-8 -*-
#本程式可將CSCC中廠商所使用的接頭料號結合進Connector_List中
import openpyxl
import re

find=0          #一般模式計數器
fuzzy=0         #模糊模式計數器

#=====================自訂函數======================
def cscc(dit, i, sheetname): #此函數作用為讀入欲創建的字典,且讀入Excel物件, 並回傳已填妥的字典檔
    if sheetname.cell(row=i,column=1).value not in cscc_conn_name: #如果字典裡尚未有此鍵,就建立此鍵並填入其值
        cscc_conn_name[sheetname.cell(row=i, column=1).value]=[sheetname.cell(row=i,column=3).value.strip() +'-'+ sheetname.cell(row=i,column=2).value.strip()]
    else:                           #如果已經有相同的鍵,就改為添加已有鍵的值
        cscc_conn_name[sheetname.cell(row=i,column=1).value].append(sheetname.cell(row=i, column=3).value.strip() +'-'+ sheetname.cell(row=i,column=2).value.strip())
        #cscc_conn_name[sheetname.cell(row=i,column=1).value].append({sheetname.cell(row=i, column=3).value : sheetname.cell(row=i, column=5).value})
    return dit

def loadExcel(filename): #此函數作用為創建一個EXCEL物件,可用來處理EXCEL的欄位
    wb=openpyxl.load_workbook(filename)
    sheet=wb.get_sheet_by_name('Sheet1')
    return wb,sheet

def Ncell(sheetname,rows,columns):
    return sheetname.cell(row=rows,column=columns).value

def cutpice(connector, cscc_conn_name, sheet2 ,i): #模糊比對模式
    global fuzzy  #全域變數要在這邊宣告
    conn_split=connector.upper().strip().replace('_',' ').replace('-',' ').split()
    s1=set(conn_split)
    for cscc_name in cscc_conn_name:
        cscc_split=cscc_name.upper().strip().replace('_',' ').replace('-',' ').split() #先將底線與中線都先轉換成空格,以方便統一以空格切分
        s2=set(cscc_split)
        if len(s1.symmetric_difference(s2)) <2: #此處運用python list比對技巧,如果兩個set的值差異的項目小於2就當它們是一樣的東西
            print("The fuzzy match item is %s" % cscc_name)
            sheet2.cell(row=i, column=3).value=cscc_conn_name[cscc_name][1].upper().strip()+' '+cscc_conn_name[cscc_name][0].upper().strip()
            fuzzy += 1
            break
        else:
            sheet2.cell(row=i, column=3).value ="Can't find match item"
    

            
#=======================載入cscc================================
wb1,sheet1=loadExcel('cscc.xlsx')
cscc_conn_name={}

for i in range(2,sheet1.max_row+1):
    cscc_conn_name=cscc(cscc_conn_name, i, sheet1)
    
#=======================載入Connector_list=======================

wb2,sheet2=loadExcel('CONNCECTOR_LIST.xlsx')
for i in range(2,sheet2.max_row+1):
    #如果connector_list裡的接頭功能名稱有在cscc中
    if Ncell(sheet2,i,1) in cscc_conn_name:
        #且其list長度值大於2(代表該功能有複數顆接頭)
        if len(cscc_conn_name[Ncell(sheet2,i,1)])> 2:
            #先比對該接頭功能名稱的料號是否有出現在CSCC中, 有的話就把該值填入connector_list中(意指廠商用料與connector指定相符)
            for j in range(0,len(cscc_conn_name[Ncell(sheet2,i,1)]),2):
                if Ncell(sheet2,i,2).replace(' ','').replace('-','').upper()==(cscc_conn_name[Ncell(sheet2,i,1)][j+1]+cscc_conn_name[Ncell(sheet2,i,1)][j]).replace(' ','').replace('-','').upper():
                    sheet2.cell(row=i,column=3).value=cscc_conn_name[Ncell(sheet2,i,1)][j+1].upper().strip()+' '+cscc_conn_name[Ncell(sheet2,i,1)][j].upper().strip()
                    print("%s is matched" %  Ncell(sheet2,i,1))
                    find += 1
                #如果前面無比對出, 那就比對connector_list的接頭孔數是否與cscc中的日產編號型式相同,是的話填入connector_list中(意指廠商使用LOCAL件)
                elif Ncell(sheet2,i,4)==int(re.findall(r'\w*(\d\d)\w*',cscc_conn_name[Ncell(sheet2,i,1)][j])[0]):
                    sheet2.cell(row=i,column=3).value=cscc_conn_name[Ncell(sheet2,i,1)][j+1].upper().strip()+' '+cscc_conn_name[Ncell(sheet2,i,1)][j].upper().strip()
                    print("%s is matched" %  Ncell(sheet2,i,1))
                    find += 1
        else:  #如果list長度沒大於2,那就直接將值填入connector_list(代表該功能cscc與connector_list是一對一的關係)
            sheet2.cell(row=i,column=3).value=cscc_conn_name[Ncell(sheet2,i,1)][1].upper().strip()+' '+cscc_conn_name[Ncell(sheet2,i,1)][0].upper().strip()
            print("%s is matched" %  Ncell(sheet2,i,1))
            find += 1
    else: #如果接頭名稱不在CSCC中,進入模糊比對模式
        cutpice(Ncell(sheet2,i,1), cscc_conn_name, sheet2, i)

maxConnector=sheet2.max_row-1
print("There are %s connector" % maxConnector)
print("Find %s match item" % find)
print("Find %s Fuzzy match item" % fuzzy)
print("Save File...")
wb2.save('CONNECTOR_LIST_Test.xlsx')
print("Done")
    


            
    
