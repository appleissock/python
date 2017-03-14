import openpyxl
from openpyxl.styles import PatternFill

wb=openpyxl.load_workbook('excel.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
whiteFill = PatternFill(fgColor='DC143C', fill_type='darkUp')
sheet.cell(row=1, column=1).fill=whiteFill
sheet.cell(row=2,column=2).fill=PatternFill(fgColor='DC143C', fill_type="lightHorizontal")

sheet['A3']='TEST'

sheet.cell(row=4, column=1).value='YAZAKI'
for i in range(10):
     name= i
     sheet.cell(row=3+i, column=1, value="%s" % name).fill=whiteFill
print("Done")

wb.save('excel_test.xlsx')
